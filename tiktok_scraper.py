#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TikTok Video Scraper
Scrapes video links and view counts from TikTok profile pages and exports to Excel
"""

import os
import time
import re
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side





class TikTokScraper:
    def __init__(self):
        self.driver = None
        self.videos_data = []
        
    def setup_chrome(self):
        """Connect to Chrome you already opened"""
        print("� Kết nối Chrome...")
        
        import os
        from selenium import webdriver
        from selenium.webdriver.chrome.service import Service
        from selenium.webdriver.chrome.options import Options
        
        # Check chromedriver
        chromedriver_path = os.path.join(os.path.dirname(__file__), "chromedriver.exe")
        if not os.path.exists(chromedriver_path):
            print("❌ Thiếu chromedriver.exe!")
            print("Chạy: python download_chromedriver.py")
            raise FileNotFoundError("chromedriver.exe")
        
        # Connect to Chrome on debugging port
        options = Options()
        options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
        
        try:
            service = Service(chromedriver_path)
            self.driver = webdriver.Chrome(service=service, options=options)
            print("✅ Kết nối thành công!")
        except Exception as e:
            print(f"\n❌ Lỗi: {str(e)}")
            print("\n� BẠN PHẢI:")
            print("1. Chạy OPEN_CHROME.bat TRƯỚC")
            print("2. Đợi Chrome mở xong")
            print("3. MỚI chạy RUN_ALL.bat")
            raise
    
    def convert_view_count(self, view_text):
        """Convert view count text to number (e.g., '31.4K' -> 31400)"""
        view_text = view_text.strip().upper()
        
        # Remove any non-numeric characters except K, M, B, and decimal point
        view_text = re.sub(r'[^\d.KMB]', '', view_text)
        
        multiplier = 1
        if 'K' in view_text:
            multiplier = 1000
            view_text = view_text.replace('K', '')
        elif 'M' in view_text:
            multiplier = 1000000
            view_text = view_text.replace('M', '')
        elif 'B' in view_text:
            multiplier = 1000000000
            view_text = view_text.replace('B', '')
        
        try:
            number = float(view_text)
            return int(number * multiplier)
        except:
            return 0
    
    def scroll_to_load_all_videos(self):
        """Scroll down to load all videos dynamically"""
        print("📜 Đang scroll để tải tất cả video...")
        
        last_height = self.driver.execute_script("return document.body.scrollHeight")
        scroll_pause_time = 2
        no_change_count = 0
        
        while True:
            # Scroll down
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(scroll_pause_time)
            
            # Calculate new scroll height
            new_height = self.driver.execute_script("return document.body.scrollHeight")
            
            if new_height == last_height:
                no_change_count += 1
                if no_change_count >= 3:  # Stop after 3 consecutive no-changes
                    break
            else:
                no_change_count = 0
                
            last_height = new_height
            
        print("✅ Đã tải xong tất cả video!")
    
    def scrape_profile(self, profile_url):
        """Scrape videos from TikTok profile"""
        print(f"🎯 Đang truy cập: {profile_url}")
        
        self.driver.get(profile_url)
        time.sleep(3)
        
        # Wait for videos to load
        try:
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "[data-e2e='user-post-item']"))
            )
        except Exception as e:
            print(f"⚠️ Không tìm thấy video. Có thể cần đăng nhập hoặc URL không đúng.")
            return
        
        # Scroll to load all videos
        self.scroll_to_load_all_videos()
        
        # Find all video elements
        print("🔍 Đang thu thập thông tin video...")
        
        video_elements = self.driver.find_elements(By.CSS_SELECTOR, "[data-e2e='user-post-item']")
        
        for idx, video_elem in enumerate(video_elements, 1):
            try:
                # Get video link
                link_elem = video_elem.find_element(By.TAG_NAME, "a")
                video_link = link_elem.get_attribute("href")
                
                # Get view count - try multiple selectors
                view_count = 0
                try:
                    # Try to find view count element
                    view_elem = video_elem.find_element(By.CSS_SELECTOR, "strong[data-e2e='video-views']")
                    view_text = view_elem.text
                    view_count = self.convert_view_count(view_text)
                except:
                    try:
                        # Alternative selector
                        view_elem = video_elem.find_element(By.CSS_SELECTOR, "strong")
                        view_text = view_elem.text
                        view_count = self.convert_view_count(view_text)
                    except:
                        pass
                
                self.videos_data.append({
                    'stt': idx,
                    'link': video_link,
                    'views': view_count
                })
                
                print(f"  [{idx}] {video_link} - {view_count:,} views")
                
            except Exception as e:
                print(f"  ⚠️ Lỗi khi xử lý video {idx}: {str(e)}")
                continue
        
        print(f"✅ Đã thu thập {len(self.videos_data)} video!")
    
    def export_to_excel(self):
        """Export scraped data to Excel"""
        if not self.videos_data:
            print("⚠️ Không có dữ liệu để xuất!")
            return
        
        print("📊 Đang tạo file Excel...")
        
        # Create output directory
        output_dir = "output"
        os.makedirs(output_dir, exist_ok=True)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "TikTok Videos"
        
        # Define border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        headers = ['STT', 'Link Video', 'Số Lượt Xem']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # Write data
        for video in self.videos_data:
            row = video['stt'] + 1
            
            # STT
            cell = ws.cell(row=row, column=1, value=video['stt'])
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            
            # Link
            cell = ws.cell(row=row, column=2, value=video['link'])
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border
            
            # Views
            cell = ws.cell(row=row, column=3, value=video['views'])
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = thin_border
            cell.number_format = '#,##0'
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 18
        
        # Save file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"tiktok_videos_{timestamp}.xlsx"
        filepath = os.path.join(output_dir, filename)
        wb.save(filepath)
        
        print(f"✅ Đã lưu file: {filepath}")
        print(f"📁 Tổng số video: {len(self.videos_data)}")
        
        return filepath
    
    def run(self):
        """Main execution flow"""
        try:
            print("\n" + "="*60)
            print("🎬 TIKTOK VIDEO SCRAPER")
            print("="*60)
            
            # Get username or URL from user
            user_input = input("\n📌 Nhập @username hoặc URL: ").strip()
            
            # Construct URL
            if user_input.startswith('http'):
                profile_url = user_input
            elif user_input.startswith('@'):
                profile_url = f"https://www.tiktok.com/{user_input}"
            else:
                profile_url = f"https://www.tiktok.com/@{user_input}"
            
            print(f"🔗 URL: {profile_url}")
            print()
            
            # Setup Chrome
            self.setup_chrome()
            
            # Scrape videos
            self.scrape_profile(profile_url)
            
            # Export to Excel
            print()
            self.export_to_excel()
            
            print("\n✨ Hoàn thành!")
            
        except Exception as e:
            print(f"\n❌ Lỗi: {str(e)}")
            
        finally:
            if self.driver:
                print("\n⏳ Đóng trình duyệt sau 3 giây...")
                time.sleep(3)
                try:
                    self.driver.quit()
                except:
                    pass


if __name__ == "__main__":
    scraper = TikTokScraper()
    scraper.run()
